from datetime import date
from uuid import UUID

from fastapi import APIRouter, Depends, UploadFile, File
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
from sqlalchemy.orm import Session
import openpyxl
from io import BytesIO
from datetime import datetime

from app.agents.attendance_agent.tools import (
    attendance_calendar,
    attendance_dashboard,
    attendance_detail,
    attendance_matrix,
    attendance_summary,
    find_employee_or_raise,
    record_attendance,
)
from app.api.deps import get_current_user, require_permissions
from app.db.session import get_db
from app.models.auth import User
from app.models.employee import Employee

router = APIRouter()


class AttendanceActionRequest(BaseModel):
    employee_id: str
    attendance_date: date
    status: str
    remarks: str | None = None


@router.get("/matrix", dependencies=[Depends(require_permissions("attendance:view"))])
def matrix(
    month: int,
    year: int,
    employee: str | None = None,
    department: str | None = None,
    status: str | None = None,
    page: int = 1,
    page_size: int = 10,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return attendance_matrix(db, month=month, year=year, employee=employee, department=department, status=status, page=page, page_size=page_size)


@router.get("/calendar", dependencies=[Depends(require_permissions("attendance:view"))])
def calendar(month: int, year: int, employee: str | None = None, department: str | None = None, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return attendance_calendar(db, month=month, year=year, employee=employee, department=department)


@router.get("/dashboard", dependencies=[Depends(require_permissions("attendance:view"))])
def dashboard(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return attendance_dashboard(db)


@router.get("/detail", dependencies=[Depends(require_permissions("attendance:view"))])
def detail(employee_id: str, attendance_date: date, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return attendance_detail(db, employee_id=employee_id, attendance_date=attendance_date)


@router.get("/employees/{employee_id}/summary", dependencies=[Depends(require_permissions("attendance:view"))])
def employee_monthly_summary(
    employee_id: UUID,
    month: int,
    year: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    employee = db.get(Employee, employee_id)
    if not employee:
        return {}
    return attendance_summary(db, employee=employee, month=month, year=year)


@router.post("/actions", dependencies=[Depends(require_permissions("attendance:manage"))])
def action(payload: AttendanceActionRequest, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    employee = None
    try:
        employee = db.get(Employee, UUID(str(payload.employee_id)))
    except ValueError:
        employee = None
    employee = employee or find_employee_or_raise(db, payload.employee_id)
    record = record_attendance(
        db,
        employee=employee,
        attendance_date=payload.attendance_date,
        status=payload.status,
        remarks=payload.remarks or "Updated from Attendance Matrix",
        actor_id=current_user.id,
        action="attendance.corrected",
    )
    db.commit()
    return attendance_detail(db, employee_id=str(record.employee_id), attendance_date=payload.attendance_date)


@router.get("/export", dependencies=[Depends(require_permissions("attendance:view"))])
def export_attendance_matrix(
    month: int,
    year: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    matrix = attendance_matrix(db, month=month, year=year, page=1, page_size=1000)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance {month}-{year}"
    
    headers = ["Employee ID", "Employee Name", "Department", "Designation"]
    if matrix.get("days"):
        for day in matrix["days"]:
            headers.append(f"{day['day']} ({day['weekday']})")
    ws.append(headers)
    
    for row in matrix.get("rows", []):
        row_data = [
            row["employee_id"],
            row["employee_name"],
            row["department"],
            row["designation"],
        ]
        cells_by_date = {cell["date"]: cell["status"] for cell in row["cells"]}
        for day in matrix.get("days", []):
            status_val = cells_by_date.get(day["date"], "MISSING")
            if status_val == "PRESENT":
                status_val = "1"
            elif status_val == "ABSENT":
                status_val = "0"
            elif status_val == "WEEKEND":
                status_val = "-"
            elif status_val == "HALF_DAY":
                status_val = "0.5"
            elif status_val == "PAID_LEAVE":
                status_val = "PL"
            elif status_val == "UNPAID_LEAVE":
                status_val = "UL"
            elif status_val == "WORK_FROM_HOME":
                status_val = "WFH"
            elif status_val == "HOLIDAY":
                status_val = "H"
            row_data.append(status_val)
        ws.append(row_data)
        
    for column in ["A", "B", "C", "D"]:
        ws.column_dimensions[column].width = 25
        
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=attendance_{year}_{month}.xlsx"}
    )


@router.post("/import", dependencies=[Depends(require_permissions("attendance:manage"))])
def import_attendance_matrix(
    month: int,
    year: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    contents = file.file.read()
    wb = openpyxl.load_workbook(filename=BytesIO(contents), data_only=True)
    ws = wb.active
    
    headers = [str(cell.value) for cell in ws[1]]
    day_columns = []
    
    for idx, header in enumerate(headers):
        if idx > 3 and header and "(" in header:
            day_str = header.split(" ")[0]
            try:
                day_num = int(day_str)
                date_str = date(year, month, day_num).isoformat()
                day_columns.append((idx, date_str))
            except ValueError:
                pass
                
    updates_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
            
        employee_id = str(row[0])
        try:
            employee = db.get(Employee, UUID(employee_id))
        except ValueError:
            continue
            
        if not employee:
            continue
            
        for col_idx, date_str in day_columns:
            status = row[col_idx]
            if status is not None:
                status_str = str(status).strip().upper()
                if status_str == "1":
                    status_str = "PRESENT"
                elif status_str == "0":
                    status_str = "ABSENT"
                elif status_str == "-":
                    status_str = "WEEKEND"
                elif status_str == "0.5":
                    status_str = "HALF_DAY"
                elif status_str == "PL":
                    status_str = "PAID_LEAVE"
                elif status_str == "UL":
                    status_str = "UNPAID_LEAVE"
                elif status_str == "WFH":
                    status_str = "WORK_FROM_HOME"
                elif status_str == "H":
                    status_str = "HOLIDAY"
                
                if status_str:
                    try:
                        record_attendance(
                            db,
                            employee=employee,
                            attendance_date=date.fromisoformat(date_str),
                            status=status_str,
                            remarks="Bulk imported from Excel",
                            actor_id=current_user.id,
                            action="attendance.corrected",
                        )
                        updates_count += 1
                    except Exception:
                        pass
                    
    db.commit()
    return {"message": "Import successful", "updates_count": updates_count}
