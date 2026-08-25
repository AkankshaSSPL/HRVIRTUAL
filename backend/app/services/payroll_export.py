from __future__ import annotations

import calendar
import re
import uuid
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models.payroll import CompanySettings, PayrollRun, PayrollRunItem

class MockWorksheet:
    def __init__(self, title="Sheet"):
        self.title = title
        self.rows = []
        self.max_row = 1

    def append(self, row):
        self.rows.append(row)
        self.max_row += 1
        
    def __getitem__(self, item):
        class MockCell:
            font = None
        return [MockCell()]

class MockWorkbook:
    def __init__(self):
        self.active = MockWorksheet()
        self.sheets = [self.active]

    def create_sheet(self, title):
        ws = MockWorksheet(title)
        self.sheets.append(ws)
        return ws

    def save(self, *args):
        pass

def _make_last_row_bold(ws):
    bold_font = Font(bold=True)
    for cell in ws[ws.max_row]:
        cell.font = bold_font


# Files land here; served back via GET /payroll/export/{filename}.
# NOTE: confirm this path is writable in your deployment (Docker volume,
# Windows dev path, etc.) — adjust if your other file-output services use a
# different convention.
STORAGE_DIR = Path(__file__).resolve().parents[2] / "storage" / "payroll"
STORAGE_DIR.mkdir(parents=True, exist_ok=True)

_SAFE_FILENAME_RE = re.compile(r"^[A-Za-z0-9_\-]+\.xlsx?$")


def is_safe_filename(filename: str) -> bool:
    """Rejects path traversal / unexpected extensions before any filesystem
    access. Generated filenames all match this pattern."""
    return bool(_SAFE_FILENAME_RE.match(filename)) and ".." not in filename


def _month_label(month: int, year: int) -> str:
    return f"{calendar.month_name[month]} {year}"


def _new_filename(export_type: str, month: int, year: int) -> str:
    ext = "xls" if export_type == "bank" else "xlsx"
    return f"payroll_{export_type}_{year}{month:02d}_{uuid.uuid4().hex[:8]}.{ext}"


def _load_items_with_employees(db: Session, payroll_run: PayrollRun) -> list[PayrollRunItem]:
    return list(
        db.scalars(
            select(PayrollRunItem)
            .where(PayrollRunItem.payroll_run_id == payroll_run.id, PayrollRunItem.deleted_at.is_(None))
            .options(selectinload(PayrollRunItem.employee))
        )
    )


def _employee_name(item: PayrollRunItem) -> str:
    emp = item.employee
    if not emp:
        return str(item.employee_id)
    name = f"{emp.first_name or ''} {emp.last_name or ''}".strip()
    return name or emp.employee_code or str(emp.id)


def generate_employee_sheet(db: Session, payroll_run: PayrollRun, company: CompanySettings, is_preview: bool = False) -> str | dict:
    """FULL_TIME employees — component-level breakdown from breakdown_json."""
    items = [i for i in _load_items_with_employees(db, payroll_run) if (i.breakdown_json or {}).get("employment_type") == "FULL_TIME"]

    wb = MockWorkbook() if is_preview else Workbook()
    ws = wb.active
    ws.title = "Employee Salary Sheet"
    ws.append([company.company_name])
    _make_last_row_bold(ws)
    month_str = calendar.month_name[payroll_run.month]
    year_str = str(payroll_run.year)
    ws.append([f"Employee sheet month of {month_str} {year_str}"])
    _make_last_row_bold(ws)
    ws.append([])
    
    headers = [
        "Sr.No", "Employee Name", "Days/Worked", "CTC", "BASIC", "HRA", "C.A", "EDU.A.", "MED. A", 
        "Employer PF", "WAGES", "E.P.F.12 % ON WAGES", "VPF", f"{month_str} arrears PAY".upper(), "Extra Pay", 
        "P.T.", "TDS", "Insu Pre", "Sponsored and self Ded", "Advance Ded", "Total Ded", "Net", 
        "Previous salary", "remark Hiked by", "Payable as updated Tax", "Deduction Pending", "TAX note"
    ]
    ws.append(headers)
    _make_last_row_bold(ws)

    totals = { "EmployerPF": 0, "WAGES": 0, "EPF": 0, "VPF": 0, "ExtraPay": 0, "PT": 0, "AdvanceDed": 0, "TotalDed": 0, "Net": 0 }

    for idx, item in enumerate(items, 1):
        bd = item.breakdown_json or {}
        earnings = bd.get("earnings", {})
        deductions = bd.get("statutory_deductions", {})
        employer = bd.get("employer_contributions", {})
        
        epf_val = deductions.get("EPF", 0)
        # Attempt to reverse calculate EPF Wages
        wages = round(epf_val / 0.12) if epf_val else 0
        
        ws.append([
            idx,
            _employee_name(item),
            bd.get("days_worked", ""),
            bd.get("gross_salary", 0),
            earnings.get("BASIC", 0),
            earnings.get("HRA", 0),
            earnings.get("CA", 0),
            earnings.get("EA", 0),
            earnings.get("MA", 0),
            employer.get("EMPLOYER_PF", 0),
            wages,
            epf_val,
            0, # VPF
            0, # arrears
            0, # extra pay
            deductions.get("PROFESSIONAL_TAX", 0),
            deductions.get("TDS", 0),
            0, # insu pre
            0, # sponsored
            0, # advance ded
            bd.get("total_deductions", 0),
            bd.get("net_salary", 0),
            "", "", "", "", ""
        ])
        totals["EmployerPF"] += employer.get("EMPLOYER_PF", 0)
        totals["WAGES"] += wages
        totals["EPF"] += epf_val
        totals["VPF"] += 0
        totals["ExtraPay"] += 0
        totals["PT"] += deductions.get("PROFESSIONAL_TAX", 0)
        totals["AdvanceDed"] += 0
        totals["TotalDed"] += bd.get("total_deductions", 0)
        totals["Net"] += bd.get("net_salary", 0)

    # Total Row
    ws.append([
        "Total", "", "", 
        "", "", "", "", "", "",
        totals["EmployerPF"], totals["WAGES"], totals["EPF"], totals["VPF"], "", totals["ExtraPay"],
        totals["PT"], "", "", "", totals["AdvanceDed"], totals["TotalDed"], totals["Net"],
        "", "", "", "", ""
    ])

    if is_preview:
        return {"tabs": [{"name": ws.title, "rows": ws.rows} for ws in wb.sheets]}

    filename = _new_filename("employee", payroll_run.month, payroll_run.year)
    wb.save(STORAGE_DIR / filename)
    return filename


def generate_consultant_sheet(db: Session, payroll_run: PayrollRun, company: CompanySettings, is_preview: bool = False) -> str | dict:
    """CONSULTANT employees — fee/leave/TDS breakdown from breakdown_json."""
    items = [i for i in _load_items_with_employees(db, payroll_run) if (i.breakdown_json or {}).get("employment_type") == "CONSULTANT"]

    wb = MockWorkbook() if is_preview else Workbook()
    ws = wb.active
    ws.title = "Consultant Sheet"
    
    month_str = calendar.month_name[payroll_run.month]
    year_str = str(payroll_run.year)
    
    ws.append([company.company_name])
    _make_last_row_bold(ws)
    ws.append([f"Consultant Sheet — {month_str} {year_str}"])
    _make_last_row_bold(ws)
    ws.append([])
    
    headers = [
        "SR NO", "CONSULTANT'S NAME", f"{month_str} PAY", "Worked Days", f"{month_str[:3].upper()} arrears PAY", 
        "LEAVE DEDUCTION", "Advance Deduction", "EXTRA WORKING PAYMENT", "INSURANCE PRIMIUM", 
        "ACTUAL PAY", "SGST 9%", "CGST 9%", "TDS", "Insurance TDS Deduction", "NET PAY", 
        "Previous salary", "remark Hiked by"
    ]
    ws.append(headers)
    _make_last_row_bold(ws)

    for idx, item in enumerate(items, 1):
        bd = item.breakdown_json or {}
        
        monthly_fee = bd.get("monthly_fee") if "monthly_fee" in bd else bd.get("gross_salary", 0)
        leave_deduction = bd.get("leave_deduction") if "leave_deduction" in bd else bd.get("statutory_deductions", {}).get("LEAVE_DEDUCTION", 0)
        actual_pay = bd.get("actual_pay") if "actual_pay" in bd else bd.get("gross_earnings", 0)
        tds = bd.get("tds") if "tds" in bd else bd.get("statutory_deductions", {}).get("FLAT_TDS", 0)

        ws.append([
            idx,
            _employee_name(item),
            monthly_fee,
            bd.get("days_worked", ""),
            0, # arrears PAY
            leave_deduction,
            0, # advance ded
            bd.get("extra_working_pay", 0),
            0, # insurance premium
            actual_pay,
            0, # SGST
            0, # CGST
            tds,
            0, # Insurance TDS
            bd.get("net_salary", 0),
            "", ""
        ])

    if is_preview:
        return {"tabs": [{"name": ws.title, "rows": ws.rows} for ws in wb.sheets]}

    filename = _new_filename("consultant", payroll_run.month, payroll_run.year)
    wb.save(STORAGE_DIR / filename)
    return filename


def generate_bank_sheet(db: Session, payroll_run: PayrollRun, company: CompanySettings, is_preview: bool = False) -> str | dict:
    """Bank upload sheet — Excel 97-2003 (.xls) format via xlwt."""
    import xlwt

    items = _load_items_with_employees(db, payroll_run)

    if is_preview:
        # Preview mode still uses the mock approach
        wb = MockWorkbook()
        ws = wb.active
        ws.title = "Bank Sheet"

        headers = [
            "Debit account", "Beneficiary Ac No", "Beneficiary Name", "Amt", "Pay Mod",
            "Date of Payment", "IFSC", "Payable Lo", "Print Loca", "Bene Mobile No.",
            "Bene Ema", "Bene add1", "Bene add2", "Bene add3", "Bene add4",
            "Add Detai1", "Add Detai2", "Add Detai3", "Add Detai4", "Add Detai5", "Remarks"
        ]
        ws.append(headers)

        month_str = calendar.month_name[payroll_run.month]
        year_str = str(payroll_run.year)
        default_remark = f"Salary For {month_str} {year_str}"
        debit_account = company.payroll_bank_account or ""

        for item in items:
            ws.append([
                debit_account,
                item.bank_account_number or "",
                _employee_name(item),
                item.net_salary,
                "N",
                "",
                item.ifsc_code or "",
                "", "",
                item.employee.phone if item.employee else "",
                "", "", "", "", "",
                "", "", "", "", "",
                default_remark
            ])

        return {"tabs": [{"name": ws.title, "rows": ws.rows} for ws in wb.sheets]}

    # --- Real export: xlwt (.xls) ---
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Bank Sheet")

    bold_style = xlwt.easyxf("font: bold on")

    headers = [
        "Debit account", "Beneficiary Ac No", "Beneficiary Name", "Amt", "Pay Mod",
        "Date of Payment", "IFSC", "Payable Lo", "Print Loca", "Bene Mobile No.",
        "Bene Ema", "Bene add1", "Bene add2", "Bene add3", "Bene add4",
        "Add Detai1", "Add Detai2", "Add Detai3", "Add Detai4", "Add Detai5", "Remarks"
    ]
    for col, h in enumerate(headers):
        ws.write(0, col, h, bold_style)

    month_str = calendar.month_name[payroll_run.month]
    year_str = str(payroll_run.year)
    default_remark = f"Salary For {month_str} {year_str}"
    debit_account = company.payroll_bank_account or ""

    for row_idx, item in enumerate(items, 1):
        row_data = [
            debit_account,
            item.bank_account_number or "",
            _employee_name(item),
            item.net_salary,
            "N",
            "",
            item.ifsc_code or "",
            "", "",
            item.employee.phone if item.employee else "",
            "", "", "", "", "",
            "", "", "", "", "",
            default_remark
        ]
        for col, val in enumerate(row_data):
            ws.write(row_idx, col, val)

    filename = _new_filename("bank", payroll_run.month, payroll_run.year)
    wb.save(str(STORAGE_DIR / filename))
    return filename


def generate_tds_sheet(db: Session, payroll_run: PayrollRun, company: CompanySettings, is_preview: bool = False) -> str | dict:
    """Two tabs — Employee TDS and Consultant TDS — matching specific formats."""
    items = _load_items_with_employees(db, payroll_run)
    employee_items = [i for i in items if (i.breakdown_json or {}).get("employment_type") == "FULL_TIME"]
    consultant_items = [i for i in items if (i.breakdown_json or {}).get("employment_type") == "CONSULTANT"]

    wb = MockWorkbook() if is_preview else Workbook()
    
    month_str = calendar.month_name[payroll_run.month]
    year_str = str(payroll_run.year)
    
    # --- 1. Employee TDS Tab ---
    ws1 = wb.active
    ws1.title = "Employee TDS"
    ws1.append([f"TDS Details Of Employee {month_str} {year_str}"])
    _make_last_row_bold(ws1)
    ws1.append([])
    ws1.append([
        "SR. NO.", "NAME OF THE DIRECTORS & EMPLOYEE", "GROSS PAY", "EXTRA WORKING DAYS", "EXTRA WORKING Pay", 
        "TDS Difference Ded", f"{month_str} arrears PAY", "Extra Pay", "Prof.Tax.", "P.F.", "ADV DED", "Insu Pre", 
        "EPF", "LEAVE", "TDS", "Net PAY"
    ])
    _make_last_row_bold(ws1)
    
    emp_tds_total = 0
    for idx, item in enumerate(employee_items, 1):
        bd = item.breakdown_json or {}
        deductions = bd.get("statutory_deductions", {})
        employer = bd.get("employer_contributions", {})
        tds_val = deductions.get("TDS", 0)
        emp_tds_total += tds_val
        ws1.append([
            idx,
            _employee_name(item),
            bd.get("gross_salary", 0),
            0, 0, 0, 0, 0, # extra/arrears
            deductions.get("PROFESSIONAL_TAX", 0),
            employer.get("EMPLOYER_PF", 0),
            0, 0, # adv/insu
            deductions.get("EPF", 0),
            0, # leave
            tds_val,
            bd.get("net_salary", 0)
        ])
    
    ws1.append([])
    ws1.append(["", f"Total TDS of Employee =", f"₹ {emp_tds_total}"])
    
    # --- 2. Consultant TDS Tab ---
    ws2 = wb.create_sheet("Consultant TDS")
    next_month = (payroll_run.month % 12) + 1
    next_year = payroll_run.year + (1 if next_month == 1 else 0)
    next_month_str = calendar.month_name[next_month]
    
    ws2.append([f"CONSULTANCY FEES {month_str} {year_str}, PAID IN THE MONTH of {next_month_str} {next_year}"])
    _make_last_row_bold(ws2)
    ws2.append([])
    ws2.append([
        "Sr. No", "CONSULTANTS' NAME", f"{month_str} PAY", "Previous Pay", "CGST 9%", "SGST 9%", 
        "arrears PAY", f"{month_str} TDS", "NET PAY", "Remark"
    ])
    _make_last_row_bold(ws2)
    
    cons_tds_total = 0
    for idx, item in enumerate(consultant_items, 1):
        bd = item.breakdown_json or {}
        tds_val = bd.get("tds") if "tds" in bd else bd.get("statutory_deductions", {}).get("FLAT_TDS", 0)
        cons_tds_total += tds_val
        ws2.append([
            idx,
            _employee_name(item),
            bd.get("actual_pay") if "actual_pay" in bd else bd.get("gross_earnings", 0), # assuming actual pay includes leave ded
            "", # prev pay
            0, 0, 0, # cgst, sgst, arrears
            tds_val,
            bd.get("net_salary", 0),
            "" # remark
        ])
        
    ws2.append([])
    ws2.append(["", "TDS of consultant", f"₹ {cons_tds_total}"])
    ws2.append([])
    
    # Office Rent Section
    ws2.append(["", "", f"{month_str} PAY", "CGST 9%", "SGST 9%", "", "TDS", "NET PAY"])
    _make_last_row_bold(ws2)
    ws2.append(["", "Office Rent", 0, 0, 0, "", 0, 0])
    ws2.append([])
    ws2.append(["", "TDS of Consultant + TDS for Office Rent =", f"₹ {cons_tds_total}"])

    if is_preview:
        return {"tabs": [{"name": ws.title, "rows": ws.rows} for ws in wb.sheets]}

    filename = _new_filename("tds", payroll_run.month, payroll_run.year)
    wb.save(STORAGE_DIR / filename)
    return filename